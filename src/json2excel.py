#!/usr/bin/python3
# -*- coding: UTF-8 -*-

"""
@version: ??
@author: ximena
@license: MIT Licence 
@contact: xiaominghe2014@gmail.com
@file: json2excel
@time: 2018/11/13

"""
import json
import enum
from collections import OrderedDict
import openpyxl
import xlwt


class SheetStyle(enum.Enum):
    title_on_row = 0
    title_on_column = 1


class JTypeError(Exception):
    def __init__(self, err='type error'):
        Exception.__init__(self, err)


def check_json(json_dic, msg):
    if not isinstance(json_dic, dict):
        raise JTypeError(msg)


def check_array(list_data, msg):
    if not isinstance(list_data, list):
        raise TypeError(msg)


def check_null(data, msg):
    if not data:
        raise Exception(msg)


def read_json(json_file):
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f, object_pairs_hook=OrderedDict)
    return data


def json2xlsx(json_data, xlsx):
    """ convert json to xlsx file
    :param json_data: json <dict>
    :param xlsx: xlsx <file>
    :return:
    """
    check_json(json_data, '"illegal json data"')
    check_null(json_data['titles'], 'json data need titles')
    sheet_title = json_data['sheetTitle'] or 'sheetTitle'
    style = json_data['style'] or SheetStyle.title_on_row
    wb = openpyxl.Workbook()
    for sheet in wb:
        if sheet.title == sheet_title:
            ws = sheet
        else:
            ws = wb.create_sheet(title=sheet_title, index=0)
    ws.sheet_properties.tabColor = "1072BA"
    titles = json_data['titles']
    check_array(titles, "json data titles isn't a list")
    set_sheet_title(ws, titles, style)
    content = json_data['content']
    check_json(content, "json data content isn't a dict")
    set_sheet_content(ws, titles, content, style)
    wb.save(xlsx)


def set_sheet_title(sheet, titles, style):
    for row in range(len(titles)):
        if SheetStyle.title_on_row == style:
            r = row + 1
            c = 1
        elif SheetStyle.title_on_column == style:
            r = 1
            c = row + 1
        else:
            raise Exception('error SheetStyle')
        sheet.cell(row=r, column=c, value=titles[row])


def set_sheet_content(sheet, titles, content, style):
    for row in range(len(titles)):
        key = titles[row]
        content_dic = content[key]
        print(content_dic)
        if isinstance(content_dic, list):
            for i, value in enumerate(content_dic):
                if SheetStyle.title_on_row == style:
                    r = row + 1
                    c = i+2
                elif SheetStyle.title_on_column == style:
                    r = i+2
                    c = row + 1
                else:
                    raise Exception('error SheetStyle')
                sheet.cell(row=r, column=c, value=value)


def json2xls(json_data, xls):
    """ convert json to xls file
    :param json_data: json <dict>
    :param xls: xls <file>
    :return:
    """
    check_json(json_data, 'illegal json data')
    check_null(json_data['titles'], 'json data need titles')
    sheet_title = json_data['sheetTitle'] or 'sheetTitle'
    style = json_data['style'] or SheetStyle.title_on_row
    book = xlwt.Workbook()
    sheet = book.add_sheet(sheet_title, cell_overwrite_ok=True)
    titles = json_data['titles']
    check_array(titles, "json data titles isn't a list")
    for i in range(len(titles)):
        if SheetStyle.title_on_row == style:
            sheet.write(i, 0, titles[i])
        else:
            sheet.write(0, i, titles[i])
    content = json_data['content']
    check_json(content, "json data content isn't a dict")
    for row in range(len(titles)):
        key = titles[row]
        content_dic = content[key]
        print(content_dic)
        if isinstance(content_dic, list):
            for i, value in enumerate(content_dic):
                if SheetStyle.title_on_row == style:
                    r = row
                    c = i+1
                else:
                    r = i+1
                    c = row
                sheet.write(r, c, value)
    book.save(xls)


def json_to_add_height(origin_json):
    print('=======================')
    return origin_json


def test_sample():
    json_dic = {
        "style": SheetStyle.title_on_column,
        "sheetTitle": "测试卷",
        "titles": [
            "学科",
            "题号",
            "分数",
        ],
        "content": {
            "学科": [1, 3, 5, 8],
            "题号": ["🌸", "", "🐸"],
            "分数": ["", "110"],
        }
    }
    xlsx = 'test.xlsx'
    xls = 'test.xls'
    # json2xlsx json ton xlsx file
    json2xlsx(json_dic, xlsx)
    # json2xls json ton xls file
    json2xls(json_dic, xls)


def main():
    test_sample()


if __name__ == '__main__':
    main()
